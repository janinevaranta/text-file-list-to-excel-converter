import re
import tkinter as tk

from openpyxl import Workbook
from tkinter import filedialog as fd
from tkinter import messagebox as mb

class App:
    def __init__(self, sourcefile = ""):
        self.sourcefile = sourcefile
        self.title_regex = r"([\da-zA-ZäöåÄÖÅ\(\), -]*:(?!=:))"
        self.email_regex = r"([^:]*[;])"
        # The following flag is turned True, if GUI is launched.
        self.is_gui = False

    def run(self):
        try:
            raw_data = self.load_source(self.sourcefile)
            parsed_data = self.parse_source(raw_data)
            self.convert_to_excel(parsed_data)

            if self.is_gui:
                mb.showinfo(
                    title="Success!",
                    message="Excel file created successfully!"
                )
        except:
            if self.is_gui:
                mb.showerror(
                    "Error!",
                    message="There was a problem converting the file. Make sure the file is in a correct format."
                )
    
    def run_gui(self):
        """
        Run the application in a GUI.

        """
        self.is_gui = True
        window = tk.Tk()
        window.title("Text File List to Excel Converter")
        window.geometry("500x100")

        guide_lbl = tk.Label(text="Select the source file: ")
        filename_lbl = tk.Label(text=self.sourcefile)

        fileselect_btn = tk.Button(
            text="Open",
            command=lambda: self.openfile_dialogue(filename_lbl),
        )
        convert_btn = tk.Button(
            text="Generate Excel",
            command=self.run,
        )

        guide_lbl.pack()
        filename_lbl.pack()
        fileselect_btn.pack()
        convert_btn.pack()

        window.mainloop()

    def openfile_dialogue(self, label):
        """
        Open the file dialogue window and save its output to sourcefile variable as a string.
        The user is prompted to provide a file that will be converted.
        "Label" parameter is expected to be a label object that showcases
        the filename in the GUI.
        """
        filename = fd.askopenfilename(
            title="Open a File",
            initialdir="./",
            filetypes=(("Text Files", "*.txt"), ("All Files", "*.*"))
        )
        self.sourcefile = filename
        label.config(text = filename)
        

    def load_source(self, source: str) -> str:
        """
        Load and read the source file, provided as a filename string.
        Returns the loaded file data as a string.
        """
        with open(source, "r") as f:
            raw_data = f.read()
            return raw_data
    
    def parse_source(self, raw_source: str) -> list:
        """
        Clean and parse the source data. Takes the unparsed string as an input.
        Returns the parsed content as a list.
        """
        cleaned_data = raw_source.replace("\n", "").replace("\t", "")

        titles = re.findall(self.title_regex, cleaned_data)
        emails = re.findall(self.email_regex, cleaned_data)

        # Remove the extra characters from the titles and contacts.
        def parse_titles(string: str):
            return string.replace(":", "").strip()

        def parse_emails(string: str):
            return string.replace(":", "").replace(" ", "")

        parsed_titles = list(map(parse_titles, titles))
        parsed_emails = list(map(parse_emails, emails))
        
        if len(parsed_titles) <= 0:
            parsed_titles = [""]

        def parse_contacts(string: str) -> list:
            return string.split(";")

        # Create a list of contacts by first parsing the contacts as a mapping function, 
        # then combining the list of contacts to the previously parsed title.
        parsed_contacts = list(zip(parsed_titles, list(map(parse_contacts, parsed_emails))))

        return parsed_contacts

    def convert_to_excel(self, data):
        """
        Convert the input data to an Excel file.
        """
        wb = Workbook()
        ws = wb.active

        # Create the Excel sheet.
        # Emails and groups are on their own columns.
        row = 1
        for contact in data:
            for email in contact[1]:
                ws.cell(row, 1, email)
                ws.cell(row, 2, contact[0])
                row += 1

        wb.save("output.xlsx")
